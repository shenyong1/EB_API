from openpyxl import load_workbook
wangsheng
from openpyxl.utils import get_column_letter

wb = load_workbook(filename=u'D:\EB_API\TestReport\TestReport.xlsx')


SheetNames = wb.sheetnames


ws = wb.get_sheet_by_name(SheetNames[0])


rows = ws.max_row
cols = ws.max_column
for i in range(1,rows+1):
    for l in range(1,cols+1):
        col = get_column_letter(l)
        print ws.cell('%s%s'%(col,i)).value

ws.append([3,3,3])
wb.save('D:\EB_API\TestReport\TestReport.xlsx')