# -*- coding: utf-8 -*-
import sys
import os
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)

import string
import requests
from random import choice

import tablib
from nose.config import flag
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter

from _ast import Return
from _mysql import result
from Config.EB_API_Config import *

def Modify_RoomType(**self):
    payload = {"Id": self['RoomTypeId'], 
               "RoomTypeName": self['RoomTypeName'],
               "weekdayPrice": self['weekdayPrice'],
               "IsActive": True,
               "Rooms": [
            {
                "Id":self['RoomID'],
                "RoomNumber": self['RoomNumber'],
                "IsActive": self['IsActive']
            }
        ]
        }
    r = requests.request('PUT', self['url'], headers=Headers, data=json.dumps(payload))
    Modify_RoomType_data = json.loads(r.text)

    businessCode = CommonMoudle(Modify_RoomType_data['businessCode'] , 200)
    resultCode = CommonMoudle(Modify_RoomType_data['resultCode'], 200)

    if businessCode & resultCode == True:
            NewRoomTypeName = Modify_RoomType_data['data']['RoomTypeName']
            NewIsActive = Modify_RoomType_data['data']['Rooms'][0]['IsActive']
            weekdayPrice = Modify_RoomType_data['data']['weekdayPrice']        
            RoomNumber = Modify_RoomType_data['data']['Rooms'][0]['RoomNumber']
            RoomTypeId = Modify_RoomType_data['data']['Rooms'][0]['RoomTypeId']
            Result=True 
            Modify_RoomType={'NewRoomTypeName':NewRoomTypeName,
                             'NewIsActive':NewIsActive,
                             'weekdayPrice':weekdayPrice,
                             'RoomNumber':RoomNumber,
                             'RoomTypeId':RoomTypeId,
                             'Result':Result}
            return Modify_RoomType
    else:
            Modify_RoomType={'Result':False}
            return Modify_RoomType         

def Del_RoomType(**self):
    r = requests.request('DELETE', self['url']+self['RoomTypeId'], headers=Headers)
    Del_RoomType_data = json.loads(r.text)
    businessCode=CommonMoudle(Del_RoomType_data['businessCode'] ,200)
    resultCode=CommonMoudle(Del_RoomType_data['resultCode'] ,200) 
    
    
    if businessCode & resultCode ==True:
        Del_RoomType={'Result':True}
        return Del_RoomType
    else:
        Del_RoomType={'Result':False}
        return Del_RoomType   

def Add_RoomType(**self):
        payload = {
                   "RoomTypeName": self['RoomTypeName'],
                   "weekdayPrice": self['weekdayPrice'],
                   "IsActive": True,
                   "Rooms": [
                    {
                    "RoomNumber": self['RoomNumber'],
                    "IsActive": True
                    }
                             ]


                             }
        r = requests.request('POST',
                              self['url'], 
                              headers=Headers ,
                              data=json.dumps(payload))

        Add_RoomType_data = json.loads(r.text)
        businessCode=CommonMoudle(Add_RoomType_data['businessCode'] ,200)
        resultCode=CommonMoudle(Add_RoomType_data['resultCode'] ,200) 
        Message=Add_RoomType_data['Message']
        
        if businessCode & resultCode ==True:
            RoomTypeName=Add_RoomType_data['data']['RoomTypeName']
            RoomNumber=Add_RoomType_data['data']['Rooms'][0]['RoomNumber']
            RoomTypeId=Add_RoomType_data['data']['Rooms'][0]['RoomTypeId']
            RoomID=Add_RoomType_data['data']['Rooms'][0]['Id']
            
            RoomType={'RoomTypeName':RoomTypeName,
               'RoomNumber':RoomNumber,
               'RoomTypeId':RoomTypeId,
               'RoomID':RoomID,
               'Result':True}
            return RoomType

        else:
            RoomType={'Result':False}
            return RoomType

def Search_RoomType(**self):
    r = requests.request('GET', self['url'], headers=Headers)
    Search_RoomType_data = json.loads(r.text)

    businessCode=CommonMoudle(Search_RoomType_data['businessCode'] ,200)
    resultCode=CommonMoudle(Search_RoomType_data['resultCode'] ,200) 

    if businessCode & resultCode == True:
        for v in Search_RoomType_data['data']:
            i=v
            if i['RoomTypeName']==self['RoomTypeName']:
                Search_RoomType = {"Result":True}
                return Search_RoomType
                break        
        else:
            Search_RoomType = {"Result":False}
            return Search_RoomType

    else:
        Search_RoomType = {"Result":False}
        return Search_RoomType
        
def RoomType_Status(**self):
    r = requests.request('GET', self['url']+self['RoomTypeId'], headers=Headers)
    RoomType_Status_data = json.loads(r.text)
    businessCode=CommonMoudle(RoomType_Status_data['businessCode'] ,200)
    resultCode=CommonMoudle(RoomType_Status_data['resultCode'] ,200) 

    if businessCode & resultCode ==True:
        Result=CommonMoudle(RoomType_Status_data['data'] ,True)
        RoomType_Status={'Result':Result}
        return RoomType_Status
    else:
        RoomType_Status={'Result':False}
        return RoomType_Status


def BatchAdd_RoomType(**self):
        payload = [{
                    "RoomTypeName":   self['RoomTypeName1'],
                    "weekdayPrice":   self['weekdayPrice'],
                    "OTARoomTypeName":self['OTARoomTypeName'],
                    "OTARoomTypeId":  self['OTARoomTypeId'],
                    "Rooms": [
            {
                "RoomNumber": self['RoomNumber1_1']
            },
            {
                "RoomNumber": self['RoomNumber1_2']
            }
                              ]
                   },
                {
                    "RoomTypeName":   self['RoomTypeName2'],
                    "weekdayPrice":   self['weekdayPrice'],
                    "Rooms": [
            {
                "RoomNumber": self['RoomNumber2_1']
            },
            {
                "RoomNumber": self['RoomNumber2_2']
            }
        ]
        }]
        
                   
        r = requests.request('POST', self['url'], headers=Headers ,data=json.dumps(payload))
        T=r.elapsed.microseconds/1000

        BatchAdd_RoomType_data = json.loads(r.text)
        businessCode=CommonMoudle(BatchAdd_RoomType_data['businessCode'] ,200)
        resultCode=CommonMoudle(BatchAdd_RoomType_data['resultCode'] ,200)
        Message=BatchAdd_RoomType_data['Message']

         
        if businessCode & resultCode ==True:
            BatchAdd_RoomType={'Result':True}
            return BatchAdd_RoomType

        else:
            BatchAdd_RoomType={'Result':False}
            return BatchAdd_RoomType       

        
def Save(**self):
    ft = Font(name='Arial Unicode MS',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000')

    D={'name':self['name'],
       'result':self['result'],
       'businessCode':self['businessCode'],
       'resultCode':self['resultCode'],
       'Message':self['Message'],
       'CaseNumber':self['CaseNumber']}
    wb = Workbook()
    ws = wb.active

#     ws.append((D['name'],D['result']))
#     print D['result']
#     wb.save("D:\EB_API\TestReport\TestReport.xlsx")
    ws['A1'].font = 'Date'
    ws['B1'].font = 'API Name'
    ws['C1'].font = 'Result'
    wb = load_workbook(filename=u'D:\EB_API\TestReport\TestReport.xlsx')

 
 
    SheetNames = wb.sheetnames

 
 
    ws = wb.get_sheet_by_name(SheetNames[0])
 
 
    rows = ws.max_row
    cols = ws.max_column
    for i in range(1,rows+1):
        for l in range(1,cols+1):
            col = get_column_letter(l)
#             print ws.cell('%s%s'%(col,i)).value
            
 
    ws.append((D['CaseNumber'],FileName,D['name'],D['businessCode'],D['resultCode'],D['Message'],D['result']))
    wb.save('D:\EB_API\TestReport\TestReport.xlsx')

        
                                   
if __name__ == "__main__":
    RoomTypeName=GetNumber(8)
    RoomTypeName1=GetNumber(8)
    RoomTypeName2=GetNumber(8)  
      
    RoomNumber=GetNumber(10)
    RoomNumber1=GetNumber(10)
    RoomNumber2=GetNumber(10)
    RoomNumber3=GetNumber(10)
    RoomNumber4=GetNumber(10)
    
    NewRoomTypeName=GetNumber(9)
    
    RoomType=Add_RoomType(CaseNumber='TCS_001',
                          url=RoomType_API_url,
                          RoomTypeName=RoomTypeName,
                          RoomNumber=RoomNumber,
                          weekdayPrice='300')
# 
#      
    RoomType_Status(url=RoomType_Status_url,
                    RoomTypeId=RoomType['RoomTypeId'],
                    CaseNumber='TCS_001')

#def Modify_RoomType(url,RoomTypeName,RoomTypeId,weekdayPrice,RoomNumber1,RoomID,IsActive1):
#     

#     Modify_RoomType(
#                     CaseNumber='TCS_001',
#                     url=RoomType_API_url,
#                     RoomTypeName=NewRoomTypeName,
#                     RoomTypeId=RoomType['RoomTypeId'],
#                     weekdayPrice=999,
#                     RoomNumber=RoomType['RoomNumber'],
#                     RoomID=RoomType['RoomID'],
#                     IsActive=False)
#        
#     Del_RoomType(url=RoomType_API_url, RoomTypeId=RoomType['RoomTypeId']) 
#          
#     RoomType_Status(CaseNumber='TCS_001',
#                     url=RoomType_Status_url,
#                     RoomTypeId=RoomType['RoomTypeId'])
#      
    # Search_RoomType(CaseNumber='TCS_001',
    #                 url=Search_RoomType_url,
    #                 RoomTypeName=RoomType['RoomTypeName'],
    #                 RoomNumber=RoomType['RoomNumber'],
    #                 RoomTypeId=RoomType['RoomTypeId'])
    
#     RoomType=BatchAdd_RoomType(CaseNumber='TCS_001',
#                           url=Search_RoomType_url,
#                           RoomTypeName1=RoomTypeName1,
#                           OTARoomTypeName='Test',
#                           OTARoomTypeId='Z23582JSC',
#                           RoomNumber1_1=RoomNumber1,
#                           RoomNumber1_2=RoomNumber2,
#                           RoomTypeName2=RoomTypeName2,
#                           RoomNumber2_1=RoomNumber3,
#                           RoomNumber2_2=RoomNumber4,
#                           weekdayPrice='300')
